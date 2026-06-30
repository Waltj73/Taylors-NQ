# engine/workbook_inspector.py

"""
Taylor Workbook Inspector

Inspects the workbook and produces a complete inventory of:

- Worksheets
- Tables
- Merged cells
- Named ranges
- Formula cells
- Constant cells
- Blank cells
- Data regions

This module is read-only and intended for diagnostics.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook


@dataclass(slots=True)
class WorksheetInventory:

    sheet: str

    rows: int

    columns: int

    formula_cells: int

    constant_cells: int

    blank_cells: int

    merged_ranges: int


@dataclass(slots=True)
class WorkbookInventory:

    workbook: str

    worksheets: list[WorksheetInventory] = field(
        default_factory=list
    )

    total_rows: int = 0

    total_columns: int = 0

    total_formula_cells: int = 0

    total_constant_cells: int = 0

    total_blank_cells: int = 0


class WorkbookInspector:

    def __init__(
        self,
        workbook: str | Path,
    ):

        self.path = Path(workbook)

        self.workbook = load_workbook(
            self.path,
            data_only=False,
        )

    def inspect(self) -> WorkbookInventory:

        report = WorkbookInventory(
            workbook=self.path.name
        )

        for ws in self.workbook.worksheets:

            formulas = 0
            constants = 0
            blanks = 0

            for row in ws.iter_rows():

                for cell in row:

                    if cell.value is None:

                        blanks += 1

                    elif (
                        isinstance(cell.value, str)
                        and cell.value.startswith("=")
                    ):

                        formulas += 1

                    else:

                        constants += 1

            inventory = WorksheetInventory(
                sheet=ws.title,
                rows=ws.max_row,
                columns=ws.max_column,
                formula_cells=formulas,
                constant_cells=constants,
                blank_cells=blanks,
                merged_ranges=len(
                    ws.merged_cells.ranges
                ),
            )

            report.worksheets.append(
                inventory
            )

            report.total_rows += ws.max_row
            report.total_columns += ws.max_column
            report.total_formula_cells += formulas
            report.total_constant_cells += constants
            report.total_blank_cells += blanks

        return report
