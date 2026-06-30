# engine/formula_registry.py

"""
Taylor Formula Registry

Builds a registry of every formula contained in the Taylor workbook.

This module is workbook-driven. It allows the application to discover
formulas instead of relying on hard-coded column assumptions.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass(slots=True)
class FormulaCell:

    sheet: str

    address: str

    formula: str


class FormulaRegistry:

    def __init__(
        self,
        workbook: str | Path,
    ):

        self.path = Path(workbook)

        self.workbook = load_workbook(
            self.path,
            data_only=False,
        )

    def all(self) -> list[FormulaCell]:

        formulas = []

        for sheet in self.workbook.worksheets:

            for row in sheet.iter_rows():

                for cell in row:

                    if (
                        isinstance(cell.value, str)
                        and cell.value.startswith("=")
                    ):

                        formulas.append(
                            FormulaCell(
                                sheet=sheet.title,
                                address=cell.coordinate,
                                formula=cell.value,
                            )
                        )

        return formulas

    def sheet(
        self,
        sheet_name: str,
    ) -> list[FormulaCell]:

        ws = self.workbook[sheet_name]

        formulas = []

        for row in ws.iter_rows():

            for cell in row:

                if (
                    isinstance(cell.value, str)
                    and cell.value.startswith("=")
                ):

                    formulas.append(
                        FormulaCell(
                            sheet=sheet_name,
                            address=cell.coordinate,
                            formula=cell.value,
                        )
                    )

        return formulas

    def formula_count(self) -> int:

        return len(self.all())

    def addresses(
        self,
        sheet_name: str,
    ) -> list[str]:

        return [
            cell.address
            for cell in self.sheet(sheet_name)
        ]

    def lookup(
        self,
        sheet_name: str,
        address: str,
    ) -> str | None:

        ws = self.workbook[sheet_name]

        value = ws[address].value

        if (
            isinstance(value, str)
            and value.startswith("=")
        ):
            return value

        return None
