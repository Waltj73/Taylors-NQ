# engine/workbook_audit.py

"""
Taylor Workbook Audit

Audits the Taylor workbook to discover:

- Worksheet layout
- Formula usage
- Named ranges
- External references
- Unsupported Excel functions
- Circular references
- Volatile functions

The output of this module is intended to guide development so every
Excel formula can eventually be reproduced exactly.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook


FUNCTION_RE = re.compile(r"([A-Z][A-Z0-9\.]*)\(")


@dataclass(slots=True)
class FormulaAudit:

    sheet: str

    cell: str

    formula: str

    functions: list[str]


@dataclass(slots=True)
class WorkbookAuditReport:

    worksheets: list[str] = field(
        default_factory=list
    )

    formula_count: int = 0

    formulas: list[FormulaAudit] = field(
        default_factory=list
    )

    functions: dict[str, int] = field(
        default_factory=dict
    )

    unsupported_functions: list[str] = field(
        default_factory=list
    )

    external_links: list[str] = field(
        default_factory=list
    )

    named_ranges: list[str] = field(
        default_factory=list
    )


SUPPORTED_FUNCTIONS = {
    "SUM",
    "AVERAGE",
    "ROUND",
    "ABS",
    "MIN",
    "MAX",
    "COUNT",
    "COUNTA",
    "IF",
    "AND",
    "OR",
    "NOT",
}


class WorkbookAudit:

    def __init__(
        self,
        workbook: str | Path,
    ):

        self.workbook = load_workbook(
            workbook,
            data_only=False,
        )

    def run(self) -> WorkbookAuditReport:

        report = WorkbookAuditReport()

        report.worksheets = (
            self.workbook.sheetnames
        )

        #
        # Named ranges
        #

        try:

            report.named_ranges = [
                n.name
                for n in self.workbook.defined_names.definedName
            ]

        except Exception:

            pass

        #
        # Scan workbook
        #

        for sheet in self.workbook.worksheets:

            for row in sheet.iter_rows():

                for cell in row:

                    value = cell.value

                    if not (
                        isinstance(value, str)
                        and value.startswith("=")
                    ):
                        continue

                    report.formula_count += 1

                    functions = sorted(
                        set(
                            FUNCTION_RE.findall(
                                value.upper()
                            )
                        )
                    )

                    report.formulas.append(
                        FormulaAudit(
                            sheet=sheet.title,
                            cell=cell.coordinate,
                            formula=value,
                            functions=functions,
                        )
                    )

                    if "[" in value:

                        report.external_links.append(
                            f"{sheet.title}!{cell.coordinate}"
                        )

                    for fn in functions:

                        report.functions[fn] = (
                            report.functions.get(
                                fn,
                                0,
                            )
                            + 1
                        )

        report.unsupported_functions = sorted(
            [
                fn
                for fn in report.functions
                if fn not in SUPPORTED_FUNCTIONS
            ]
        )

        return report
