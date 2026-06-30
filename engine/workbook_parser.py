# engine/workbook_parser.py

"""
Taylor Workbook Parser

Reads the Taylor workbook structure directly from Excel and extracts
the formulas, headers, aliases, and calculation metadata.

This parser is used so the application stays synchronized with the
actual workbook instead of hard-coded assumptions.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass(slots=True)
class WorkbookColumn:

    excel_column: str

    title: str

    formula: str | None

    header_row: int


class WorkbookParser:

    def __init__(self, workbook: str | Path):

        self.path = Path(workbook)

        self.workbook = load_workbook(
            self.path,
            data_only=False,
        )

    @property
    def sheets(self):

        return self.workbook.sheetnames

    def parse_sheet(
        self,
        sheet_name: str = "NQ",
    ) -> list[WorkbookColumn]:

        ws = self.workbook[sheet_name]

        columns = []

        #
        # Row 3 contains the visible column names.
        # Row 5 contains the first formulas.
        #

        for cell in ws[3]:

            if cell.value is None:
                continue

            excel_col = cell.column_letter

            formula = ws[f"{excel_col}5"].value

            if isinstance(formula, str):

                if not formula.startswith("="):
                    formula = None

            else:

                formula = None

            columns.append(
                WorkbookColumn(
                    excel_column=excel_col,
                    title=str(cell.value).strip(),
                    formula=formula,
                    header_row=3,
                )
            )

        return columns

    def formula_map(
        self,
        sheet_name: str = "NQ",
    ) -> dict[str, str]:

        mapping = {}

        for column in self.parse_sheet(sheet_name):

            if column.formula:

                mapping[column.title] = column.formula

        return mapping

    def print_formula_summary(self):

        for column in self.parse_sheet():

            print(
                f"{column.excel_column:>3} | "
                f"{column.title:<45} | "
                f"{column.formula}"
            )
