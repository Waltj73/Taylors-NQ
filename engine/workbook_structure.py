# engine/workbook_structure.py

"""
Taylor Workbook Structure

Discovers the physical layout of the Taylor workbook.

This module is intentionally read-only. It does not calculate
anything. Its responsibility is to identify:

- Header rows
- Data start row
- Formula row
- Used columns
- Empty columns
- Worksheet dimensions
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass(slots=True)
class ColumnDefinition:

    letter: str

    header: str | None

    first_formula_row: int | None

    has_formula: bool

    has_data: bool


@dataclass(slots=True)
class WorksheetStructure:

    sheet: str

    first_row: int

    last_row: int

    first_column: int

    last_column: int

    columns: list[ColumnDefinition]


class WorkbookStructure:

    def __init__(
        self,
        workbook: str | Path,
    ):

        self.workbook = load_workbook(
            workbook,
            data_only=False,
        )

    def analyze(
        self,
        sheet_name: str,
    ) -> WorksheetStructure:

        ws = self.workbook[sheet_name]

        columns = []

        for column in ws.iter_cols():

            first_formula = None
            header = None
            has_data = False

            for cell in column:

                if header is None and cell.value not in (None, ""):
                    header = str(cell.value)

                if isinstance(cell.value, str):

                    if cell.value.startswith("="):

                        has_data = True

                        if first_formula is None:
                            first_formula = cell.row

                elif cell.value is not None:

                    has_data = True

            columns.append(
                ColumnDefinition(
                    letter=column[0].column_letter,
                    header=header,
                    first_formula_row=first_formula,
                    has_formula=first_formula is not None,
                    has_data=has_data,
                )
            )

        return WorksheetStructure(
            sheet=sheet_name,
            first_row=ws.min_row,
            last_row=ws.max_row,
            first_column=ws.min_column,
            last_column=ws.max_column,
            columns=columns,
        )

    def analyze_all(self):

        return {
            sheet: self.analyze(sheet)
            for sheet in self.workbook.sheetnames
        }
