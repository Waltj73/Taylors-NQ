# engine/workbook_loader.py

"""
Taylor Workbook Loader

Loads the Taylor workbook into an in-memory model that can be used
by the Excel runtime.

Unlike workbook.py, this loader preserves:

    • formulas
    • calculated values
    • row numbers
    • column letters

The goal is to faithfully reproduce Excel in Python.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass(slots=True)
class WorkbookCell:

    sheet: str

    row: int

    column: int

    letter: str

    coordinate: str

    value: object

    formula: str | None


class WorkbookLoader:

    def __init__(
        self,
        workbook: str | Path,
    ):

        self.workbook = load_workbook(
            workbook,
            data_only=False,
        )

    def cells(
        self,
        sheet_name: str,
    ) -> list[WorkbookCell]:

        ws = self.workbook[sheet_name]

        results = []

        for row in ws.iter_rows():

            for cell in row:

                formula = None

                if (
                    isinstance(cell.value, str)
                    and cell.value.startswith("=")
                ):
                    formula = cell.value

                results.append(
                    WorkbookCell(
                        sheet=sheet_name,
                        row=cell.row,
                        column=cell.column,
                        letter=cell.column_letter,
                        coordinate=cell.coordinate,
                        value=cell.value,
                        formula=formula,
                    )
                )

        return results

    def formulas(
        self,
        sheet_name: str,
    ):

        return [
            c
            for c in self.cells(sheet_name)
            if c.formula is not None
        ]

    def constants(
        self,
        sheet_name: str,
    ):

        return [
            c
            for c in self.cells(sheet_name)
            if c.formula is None
        ]

    def worksheet(self, sheet_name: str):

        return self.workbook[sheet_name]
