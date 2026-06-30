# engine/workbook_metadata.py

"""
Taylor Workbook Metadata

Extracts structural metadata from the Taylor workbook so the
application can automatically adapt if the workbook changes.

No calculations.
No UI.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass(slots=True)
class SheetMetadata:

    name: str

    rows: int

    columns: int

    max_row: int

    max_column: int

    hidden: bool


class WorkbookMetadata:

    def __init__(
        self,
        workbook: str | Path,
    ):

        self.path = Path(workbook)

        self.workbook = load_workbook(
            self.path,
            data_only=False,
        )

    @property
    def workbook_name(self):

        return self.path.name

    @property
    def sheet_names(self):

        return self.workbook.sheetnames

    def sheet(
        self,
        sheet_name: str,
    ) -> SheetMetadata:

        ws = self.workbook[sheet_name]

        return SheetMetadata(
            name=sheet_name,
            rows=ws.max_row,
            columns=ws.max_column,
            max_row=ws.max_row,
            max_column=ws.max_column,
            hidden=ws.sheet_state != "visible",
        )

    def all_sheets(self):

        return [
            self.sheet(name)
            for name in self.sheet_names
        ]

    def workbook_summary(self):

        return {
            "file": self.workbook_name,
            "worksheets": len(
                self.sheet_names
            ),
            "sheet_names": self.sheet_names,
        }

    def active_sheet(self):

        return self.workbook.active.title

    def has_sheet(
        self,
        name: str,
    ):

        return name in self.sheet_names

    def dimensions(
        self,
        sheet_name: str,
    ):

        ws = self.workbook[sheet_name]

        return ws.dimensions

    def used_range(
        self,
        sheet_name: str,
    ):

        ws = self.workbook[sheet_name]

        return (
            ws.min_row,
            ws.max_row,
            ws.min_column,
            ws.max_column,
        )
